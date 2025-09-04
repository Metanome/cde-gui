"""
Data processor for handling extraction rules and generating Excel output.
"""

import pandas as pd
from typing import List, Dict, Any, Optional
from ..utils.data_transformer import DataTransformer


class DataProcessor:
    """Processes extracted data according to rules and generates Excel output."""
    
    def __init__(self, extraction_rules: List[Dict[str, str]]):
        """
        Initialize data processor with extraction rules.
        
        Args:
            extraction_rules: List of extraction rule dictionaries
        """
        self.extraction_rules = extraction_rules
        self.transformer = DataTransformer()
    
    def process_extracted_data(self, raw_data: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        Process raw extracted data according to rules.
        
        Args:
            raw_data: List of raw data dictionaries from extraction
            
        Returns:
            Processed pandas DataFrame ready for Excel export
        """
        processed_data = []
        
        for record in raw_data:
            processed_record = {
                'Subject ID': record.get('subject_id', ''),
                'Name': record.get('patient_name', '')
            }
            
            # Apply extraction rules if text was successfully extracted
            if record.get('extracted_text'):
                for rule in self.extraction_rules:
                    field_name = rule.get('name', '')
                    pattern = rule.get('pattern', '')
                    transform = rule.get('transform', 'none')
                    
                    if field_name and pattern:
                        # Extract value using pattern
                        raw_value = self.transformer.extract_with_pattern(
                            record['extracted_text'], 
                            pattern
                        )
                        
                        # Apply transformation
                        if raw_value is not None:
                            processed_value = self.transformer.transform_value(
                                raw_value, 
                                transform
                            )
                            processed_record[field_name] = processed_value
                        else:
                            processed_record[field_name] = None
            
            processed_data.append(processed_record)
        
        return pd.DataFrame(processed_data)
    
    def export_to_excel(self, df: pd.DataFrame, output_path: str) -> bool:
        """
        Export processed data to Excel file.
        
        Args:
            df: Processed DataFrame
            output_path: Path for output Excel file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Create Excel writer with styling
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write main data
                df.to_excel(writer, sheet_name='Extracted Data', index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Extracted Data']
                
                # Apply formatting
                self._format_excel_worksheet(worksheet, df)
                
                # Add summary sheet
                self._add_summary_sheet(writer, df)
            
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {str(e)}")
            return False
    
    def _format_excel_worksheet(self, worksheet, df: pd.DataFrame) -> None:
        """Apply formatting to Excel worksheet."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Apply header formatting
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    def _add_summary_sheet(self, writer, df: pd.DataFrame) -> None:
        """Add summary sheet with extraction statistics."""
        # Create summary data
        total_records = len(df)
        
        # Count records with extracted data (excluding just Subject ID and Name)
        data_columns = [col for col in df.columns if col not in ['Subject ID', 'Name']]
        
        # Count successful extractions (records that have at least one non-empty data field)
        successful_extractions = 0
        for _, row in df.iterrows():
            has_data = any(pd.notna(row[col]) and str(row[col]).strip() != '' 
                          for col in data_columns)
            if has_data:
                successful_extractions += 1
        
        failed_extractions = total_records - successful_extractions
        
        summary_data = {
            'Metric': [
                'Total Records',
                'Records with Extracted Data',
                'Records with No Data',
                'Success Rate (%)',
                'Data Fields Extracted'
            ],
            'Value': [
                total_records,
                successful_extractions,
                failed_extractions,
                round((successful_extractions / total_records) * 100, 2) if total_records > 0 else 0,
                len(data_columns)
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    def validate_extraction_rules(self) -> List[str]:
        """
        Validate extraction rules for common issues.
        
        Returns:
            List of validation error messages
        """
        errors = []
        
        for i, rule in enumerate(self.extraction_rules):
            rule_num = i + 1
            
            # Check required fields
            if not rule.get('name'):
                errors.append(f"Rule {rule_num}: Missing field name")
            
            if not rule.get('pattern'):
                errors.append(f"Rule {rule_num}: Missing extraction pattern")
            
            # Test regex pattern
            try:
                import re
                re.compile(rule.get('pattern', ''))
            except re.error as e:
                errors.append(f"Rule {rule_num}: Invalid regex pattern - {str(e)}")
        
        return errors
